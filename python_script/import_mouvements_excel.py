"""
Reconstruit l'historique de roster de la saison régulière active (pooler_rosters,
roster_change_log) à partir du suivi personnel de David, excel/Mouvements_consolides.xlsx
(hors dépôt git — voir CLAUDE.md section 1 et SUIVI_PROJET.md 2026-07-31/2026-08-03).

Contexte : buildStandings() (app/lib/standings.ts) attribue les points d'un joueur selon
la fenêtre added_at→removed_at de sa ligne pooler_rosters et les transitions de statut
journalisées dans roster_change_log. Cette histoire fine n'a jamais été saisie au fil de
l'eau pour 2025-26 — seule une poignée de corrections ponctuelles existe via
/admin/historique. Ce script rejoue les ~270 mouvements du fichier Excel pour reconstruire
cette histoire automatiquement, plutôt que de la ressaisir un par un dans l'UI.

Mécanique (indépendante du texte libre de la colonne "Type", jugé peu fiable — voir le
plan de session) : on simule chronologiquement, pour chaque joueur touché, qui le détient
et à quel statut, en se basant uniquement sur les colonnes "Statut joueur acquis/cédé" et
la présence de "Echange Pooler". Le côté acquis regarde l'état simulé courant du joueur
pour décider s'il s'agit d'un changement de statut interne, d'un vrai transfert, ou d'un
ajout neuf. Le côté cédé quitte réellement le pool si son statut est "Ballotage", part
vers "Echange Pooler" si rempli, sinon reste chez le même pooler à son nouveau statut.

Portée : uniquement pooler_rosters/roster_change_log. Les colonnes de choix de repêchage
(Choix acquis/cédé, Choix Pooler) sont seulement listées dans le rapport, jamais rejouées
dans pool_draft_picks (décidé avec David le 2026-08-03).

Usage:
    python import_mouvements_excel.py            # dry-run : rapport seulement, rien écrit
    python import_mouvements_excel.py --apply     # exécute réellement (confirmation "oui")
"""

import os
import sys
import re
import difflib
from collections import defaultdict
from datetime import datetime

sys.stdout.reconfigure(encoding='utf-8')

from dotenv import load_dotenv
from supabase import create_client
import openpyxl
from unidecode import unidecode

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_PATH = os.path.normpath(os.path.join(BASE_DIR, '..', 'excel', 'Mouvements_consolides.xlsx'))
CORRECTIONS_PATH = os.path.normpath(os.path.join(BASE_DIR, '..', 'excel', 'correction_violations_alignements.txt'))
FINAL_ALIGNMENT_CORRECTIONS_PATH = os.path.normpath(os.path.join(BASE_DIR, '..', 'excel', 'correction_alignements_finaux.txt'))

POOLER_NAME_MAP = {
    'vincent': 'Vincent',
    'paule': 'Paule',
    'nicolas': 'Nicolas',
    'steve': 'Steve',
    'david': 'David',
    'jerome': 'Jérôme',
    'sebastien_fau': 'Sébastien F.',
    'sebastien_stl': 'Sébastien S.',
    'sebastien s.': 'Sébastien S.',
    'sebastien f.': 'Sébastien F.',
}

STATUS_MAP = {
    'actif': 'actif',
    'reserviste': 'reserviste',
    'recrue': 'recrue',
    'ir': 'ltir',
    'ltir': 'ltir',
    'ballotage': 'BALLOTAGE',
}

ACTIVE_LIMITS = {'forward': 12, 'defense': 6, 'goalie': 2}


class Tee:
    """Duplique stdout vers un fichier de log, comme les autres scripts du pipeline."""
    def __init__(self, path):
        self.file = open(path, 'w', encoding='utf-8')
        self.stdout = sys.stdout

    def write(self, data):
        self.stdout.write(data)
        self.file.write(data)

    def flush(self):
        self.stdout.flush()
        self.file.flush()


def norm(s):
    return unidecode(str(s or '')).lower().strip()


def connect():
    load_dotenv(os.path.join(BASE_DIR, '.env.staging'), override=True)
    url = os.environ['SUPABASE_URL']
    key = os.environ['SUPABASE_SERVICE_KEY']
    return create_client(url, key)


def fetch_all(db, table, select, filters=None):
    rows, offset, page = [], 0, 1000
    filters = filters or {}
    while True:
        q = db.table(table).select(select)
        for k, v in filters.items():
            q = q.eq(k, v)
        r = q.range(offset, offset + page - 1).execute()
        rows.extend(r.data)
        if len(r.data) < page:
            break
        offset += page
    return rows


def get_active_season(db):
    r = db.table('pool_seasons').select('id, season, saison_start_date, saison_end_date') \
        .eq('is_active', True).eq('is_playoff', False).single().execute()
    if not r.data:
        raise SystemExit('[ERREUR] Aucune saison régulière active en staging.')
    return r.data


def get_player_bucket(position):
    pos = (position or '').upper()
    if 'G' in pos:
        return 'goalie'
    if 'D' in pos:
        return 'defense'
    return 'forward'


# ---------------------------------------------------------------------------
# Lecture du fichier Excel
# ---------------------------------------------------------------------------

def load_excel_rows():
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb['Mouvements']
    headers = [c.value for c in ws[1]]
    rows = [dict(zip(headers, r)) for r in ws.iter_rows(min_row=2, values_only=True)]
    return rows


def load_roster_initial_rows():
    """Onglet 'Roster_Initial' (Pooler, Joueur, Statut) — l'alignement réel de chaque
    pooler au tout début de la saison, ajouté par David le 2026-08-03 pour remplacer
    l'hypothèse par défaut 'actif' du bootstrap par la vraie donnée. Optionnel : si
    l'onglet n'existe pas encore, la simulation retombe sur le bootstrap heuristique."""
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    if 'Roster_Initial' not in wb.sheetnames:
        return None
    ws = wb['Roster_Initial']
    headers = [c.value for c in ws[1]]
    rows = [dict(zip(headers, r)) for r in ws.iter_rows(min_row=2, values_only=True)]
    return [r for r in rows if r.get('Pooler') and r.get('Joueur')]


LEGALITY_LINE_RE = re.compile(
    r'(\S.*?) \[(.*?)\] (Trop de \w+ actifs \(\d+/\d+\)|Moins de 2 .*?) : (.*)')


def load_legality_corrections():
    """excel/correction_violations_alignements.txt (optionnel, ajouté par David le
    2026-08-04) — même format que la section 'Violations de légalité' du rapport, mais
    avec la vraie liste des joueurs actifs à chaque période au lieu de celle simulée.
    Retourne pooler -> bucket -> [(date_debut, {noms corrects}), ...] trié chronologiquement."""
    if not os.path.exists(CORRECTIONS_PATH):
        return None
    groups = defaultdict(list)
    with open(CORRECTIONS_PATH, encoding='utf-8') as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            m = LEGALITY_LINE_RE.match(line)
            if not m:
                print(f"[AVERTISSEMENT] Ligne de corrections non reconnue, ignorée : {line!r}")
                continue
            pooler_raw, label, msg, names = m.groups()
            pooler = map_pooler(pooler_raw) or pooler_raw
            bucket_m = re.search(r'Trop de (\w+)', msg)
            bucket_label_to_key = {'attaquants': 'forward', 'défenseurs': 'defense', 'gardiens': 'goalie'}
            bucket = bucket_label_to_key.get(bucket_m.group(1)) if bucket_m else None
            if bucket is None:
                continue  # 'Moins de 2 réservistes' — hors scope (aucune ligne de David ne le concerne)
            start = label.split(' -> ')[0].split(' au ')[0].replace('du ', '').strip()
            corrected_names = frozenset(n.strip() for n in names.split(',') if n.strip())
            groups[(pooler, bucket)].append((start, corrected_names))
    for key in groups:
        groups[key].sort(key=lambda t: t[0])
    return groups


def load_final_alignment_corrections():
    """excel/correction_alignements_finaux.txt (optionnel, ajouté par David le 2026-08-05) —
    même format que la section 'Alignements finaux' du rapport (indentation 2/4/6 espaces :
    pooler / position / statut+nom), mais avec la vraie composition voulue au lieu de celle
    simulée. Retourne pooler -> {nom: statut}."""
    if not os.path.exists(FINAL_ALIGNMENT_CORRECTIONS_PATH):
        return None
    target = defaultdict(dict)
    current_pooler = None
    valid_statuses = {'actif', 'reserviste', 'ltir', 'recrue'}
    with open(FINAL_ALIGNMENT_CORRECTIONS_PATH, encoding='utf-8') as f:
        for raw_line in f:
            line = raw_line.rstrip('\n')
            if not line.strip():
                continue
            indent = len(line) - len(line.lstrip(' '))
            content = line.strip()
            if indent <= 2:
                m = re.match(r'(.+?)\s*\(\d+\)\s*:$', content)
                current_pooler = (map_pooler(m.group(1)) or m.group(1)) if m else None
                continue
            if indent == 4:
                continue  # en-tête de position (Attaquants/Défenseurs/Gardiens) — ignoré
            if indent >= 6 and current_pooler:
                parts = content.split(None, 1)
                if len(parts) == 2 and parts[0] in valid_statuses:
                    target[current_pooler][parts[1].strip()] = parts[0]
    return dict(target)


def to_date_str(v):
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.strftime('%Y-%m-%d')
    s = str(v).strip()
    return s[:10] if s else None


def parse_player_name(cell):
    if not cell:
        return None
    s = str(cell).strip()
    if not s:
        return None
    # format attendu : "Prénom Nom - EQUIPE - POS,POS2" — on ne garde que le nom
    return s.split(' - ')[0].strip()


def map_pooler(cell):
    if not cell:
        return None
    key = norm(cell)
    return POOLER_NAME_MAP.get(key)


def map_status(cell):
    if not cell:
        return None
    key = norm(cell)
    return STATUS_MAP.get(key)


# ---------------------------------------------------------------------------
# Résolution des noms de joueurs
# ---------------------------------------------------------------------------

class PlayerIndex:
    def __init__(self, players):
        self.by_id = {p['id']: p for p in players}
        self.by_key = defaultdict(list)
        for p in players:
            key = norm(f"{p['first_name']} {p['last_name']}")
            self.by_key[key].append(p['id'])
        self.all_names = list(self.by_key.keys())

    def resolve(self, name):
        """Retourne (player_id, note) — player_id est None si non résolu/ambigu."""
        key = norm(name)
        candidates = self.by_key.get(key)
        if candidates and len(candidates) == 1:
            return candidates[0], None
        if candidates and len(candidates) > 1:
            return None, f"Nom ambigu ({len(candidates)} joueurs) : {name}"
        suggestion = difflib.get_close_matches(key, self.all_names, n=1, cutoff=0.8)
        if suggestion:
            return None, f"Nom introuvable : {name!r} — vouliez-vous dire {self.by_id[self.by_key[suggestion[0]][0]]['first_name']} {self.by_id[self.by_key[suggestion[0]][0]]['last_name']!r} ?"
        return None, f"Nom introuvable : {name!r} (aucune suggestion)"


# ---------------------------------------------------------------------------
# Rapport
# ---------------------------------------------------------------------------

class Report:
    def __init__(self):
        self.unresolved_names = []       # (row_idx, date, pooler, side, raw_name, note)
        self.bootstraps = []             # (pooler_name, player_id, date, note)
        self.echange_mismatches = []     # (row_idx, date, pooler, player_id, declared, actual)
        self.anomalies = []              # (row_idx, date, pooler, message)
        self.pick_rows = []              # (row_idx, date, pooler, description)
        self.legality_violations = []    # (date, pooler, message)
        self.roster_initial_issues = []  # (message) — noms non résolus / statuts invalides / doublons
        self.roster_initial_count = 0
        self.legality_corrections_applied = []  # (pooler, name, action, date)
        self.legality_correction_issues = []    # (message)
        self.final_alignment_corrections_applied = []  # (name, from_pooler, from_status, to_pooler, to_status)
        self.final_alignment_correction_issues = []     # (message)
        self.rows_processed = 0
        self.players_touched = set()
        self.poolers_touched = set()


# ---------------------------------------------------------------------------
# Simulation
# ---------------------------------------------------------------------------

class Simulation:
    def __init__(self, season_start_ts):
        self.season_start_ts = season_start_ts
        self.current_owner = {}                      # player_id -> pooler_name | None
        self.rows_by_pair = defaultdict(list)         # (pooler_name, player_id) -> [row dicts]

    def preload_initial(self, pooler, player_id, ptype):
        """Pose l'état réel de départ (onglet Roster_Initial) — appelé avant de rejouer
        les mouvements, remplace le bootstrap heuristique pour tout joueur listé."""
        row = {
            'added_at': self.season_start_ts,
            'removed_at': None,
            'player_type': ptype,
            'transitions': [(self.season_start_ts, None, ptype)],
            'bootstrap': False,
        }
        self.rows_by_pair[(pooler, player_id)].append(row)
        self.current_owner[player_id] = pooler

    def get_open_row(self, pooler, player_id):
        lst = self.rows_by_pair.get((pooler, player_id))
        if lst and lst[-1]['removed_at'] is None:
            return lst[-1]
        return None

    def ensure_owner(self, player_id, pooler, date, report, reason):
        """Si le joueur n'a jamais été vu, on suppose qu'il était déjà chez `pooler`
        depuis le début de la saison, au statut 'actif' (meilleure estimation générique —
        la très grande majorité des mouvements du fichier partent d'un joueur actif qu'on
        rétrograde). Marqué dans le rapport pour vérification manuelle."""
        if self.current_owner.get(player_id) == pooler and self.get_open_row(pooler, player_id):
            return
        if self.get_open_row(pooler, player_id) is None:
            row = {
                'added_at': self.season_start_ts,
                'removed_at': None,
                'player_type': 'actif',
                'transitions': [(self.season_start_ts, None, 'actif')],
                'bootstrap': True,
            }
            self.rows_by_pair[(pooler, player_id)].append(row)
            report.bootstraps.append((pooler, player_id, date, reason))
        self.current_owner[player_id] = pooler

    def close_row(self, pooler, player_id, date, report):
        row = self.get_open_row(pooler, player_id)
        if row is None:
            report.anomalies.append((None, date, pooler,
                f"Retrait impossible : aucune ligne ouverte pour joueur {player_id} chez {pooler}"))
            return
        row['removed_at'] = date
        if self.current_owner.get(player_id) == pooler:
            self.current_owner[player_id] = None

    def open_row(self, pooler, player_id, date, ptype):
        row = {
            'added_at': date,
            'removed_at': None,
            'player_type': ptype,
            'transitions': [(date, None, ptype)],
            'bootstrap': False,
        }
        self.rows_by_pair[(pooler, player_id)].append(row)
        self.current_owner[player_id] = pooler

    def change_type(self, pooler, player_id, date, new_type, report, reason):
        """Insère une transition à `date`. Résout le statut juste AVANT `date` en parcourant
        les transitions déjà connues (pas seulement row['player_type'], qui ne reflète que
        le statut le plus RÉCENT chronologiquement) — nécessaire pour les corrections de
        légalité appliquées après coup, à une date potentiellement antérieure à une
        transition déjà traitée (ex: Jake Neighbours corrigé au 23 octobre après que le
        vrai mouvement du 11 novembre ait déjà été simulé : row['player_type'] valait déjà
        'reserviste' à ce moment-là, faisant passer la correction pour un no-op à tort).
        Recalcule ensuite row['player_type'] comme le statut de la toute dernière transition
        chronologique, pas celui de la dernière transition insérée."""
        self.ensure_owner(player_id, pooler, date, report, reason)
        row = self.get_open_row(pooler, player_id)
        status_before = row['transitions'][0][2] if row['transitions'] else row['player_type']
        for changed_at, _old_t, new_t in sorted(row['transitions'], key=lambda t: t[0]):
            if changed_at <= date:
                status_before = new_t
            else:
                break
        if status_before == new_type:
            return
        row['transitions'].append((date, status_before, new_type))
        row['transitions'].sort(key=lambda t: t[0])
        row['player_type'] = row['transitions'][-1][2]

    def process_acquis(self, pooler, player_id, date, new_type, echange_pooler, row_idx, report):
        cur = self.current_owner.get(player_id)
        if echange_pooler and cur is not None and cur != echange_pooler and cur != pooler:
            report.echange_mismatches.append((row_idx, date, pooler, player_id, echange_pooler, cur))
        if cur is None and echange_pooler and echange_pooler != pooler:
            self.ensure_owner(player_id, echange_pooler, date, report,
                               f"Origine déduite via 'Echange Pooler' (ligne {row_idx})")
            cur = echange_pooler
        if cur == pooler:
            self.change_type(pooler, player_id, date, new_type, report, f"ligne {row_idx}")
        elif cur is not None:
            self.close_row(cur, player_id, date, report)
            self.open_row(pooler, player_id, date, new_type)
        else:
            self.open_row(pooler, player_id, date, new_type)

    def process_cede(self, pooler, player_id, date, new_type_raw, echange_pooler, row_idx, report):
        cur = self.current_owner.get(player_id)

        # Le fichier consolide d'anciens onglets par pooler (extract_mouvements.py) — un
        # même échange réel entre 2 poolers peut donc apparaître 2 fois (une ligne "acquis"
        # dans le tab du receveur, une ligne "cédé" dans le tab du donneur), à des dates
        # parfois légèrement différentes. Si l'échange a déjà été appliqué (le joueur est
        # déjà chez echange_pooler au moment de traiter cette ligne), ce n'est pas un 2e
        # retrait réel — juste la réaffirmation du même événement vu de l'autre côté.
        if echange_pooler and cur == echange_pooler and echange_pooler != pooler:
            self.change_type(echange_pooler, player_id, date,
                              new_type_raw if new_type_raw != 'BALLOTAGE' else 'actif',
                              report, f"ligne {row_idx} (jumelle d'un échange déjà appliqué)")
            return

        if cur is not None and cur != pooler and cur != echange_pooler:
            report.anomalies.append((row_idx, date, pooler,
                f"Cédé déclaré chez {pooler} mais détenu par {cur} selon la simulation (joueur {player_id}) — chronologie incohérente, à vérifier dans le fichier."))

        self.ensure_owner(player_id, pooler, date, report, f"Première mention (côté cédé, ligne {row_idx})")
        if new_type_raw == 'BALLOTAGE':
            self.close_row(pooler, player_id, date, report)
        elif echange_pooler and echange_pooler != pooler:
            self.close_row(pooler, player_id, date, report)
            self.open_row(echange_pooler, player_id, date, new_type_raw or 'actif')
        elif new_type_raw:
            self.change_type(pooler, player_id, date, new_type_raw, report, f"ligne {row_idx}")
        else:
            report.anomalies.append((row_idx, date, pooler, f"Statut cédé manquant pour joueur {player_id}"))


# ---------------------------------------------------------------------------
# Légalité (best-effort, non bloquant — cap volontairement omis, voir plan)
# ---------------------------------------------------------------------------

def get_roster_snapshot(sim, baseline_roster, touched_ids, pooler, date):
    """baseline_roster: pooler_name -> {player_id: player_type} pour les joueurs JAMAIS
    touchés par le fichier (état DB courant, valide tout au long puisque non modifié).
    Les joueurs touchés sont exclus du baseline — leur présence/statut à `date` vient
    uniquement de l'état simulé (absent si pas encore acquis ou déjà cédé à `date`)."""
    roster = {pid: t for pid, t in baseline_roster.get(pooler, {}).items() if pid not in touched_ids}
    for (p, pid), lst in sim.rows_by_pair.items():
        if p != pooler:
            continue
        row = lst[-1]
        if row['added_at'] <= date and (row['removed_at'] is None or row['removed_at'] > date):
            # Statut à `date`, pas le statut final de la ligne — cette fonction est appelée
            # après la fin de la simulation complète (corrections + vérification finale),
            # donc row['player_type'] reflète le tout dernier changement, pas forcément celui
            # en vigueur à `date` (ex: un joueur redevenu actif plus tard dans la saison
            # apparaîtrait à tort comme actif à une date antérieure où il était recrue).
            status = row['player_type']
            for changed_at, _old_type, new_type in sorted(row['transitions'], key=lambda t: t[0]):
                if changed_at <= date:
                    status = new_type
                else:
                    break
            roster[pid] = status
    return roster


def compute_final_rosters(sim, baseline_roster, all_touched_ids, positions, pindex):
    """pooler -> [(bucket, status, name, player_id), ...] — combine les joueurs simulés
    (touchés + préchargés via Roster_Initial, tous déjà dans sim.rows_by_pair) et les
    joueurs jamais touchés (statut DB actuel, inchangé). Réutilisé pour l'affichage des
    alignements finaux et pour l'application des corrections d'alignement final."""
    by_pooler = defaultdict(list)
    for (pooler, pid), lst in sim.rows_by_pair.items():
        row = lst[-1]
        if row['removed_at'] is not None:
            continue
        p = pindex.by_id.get(pid, {})
        name = f"{p.get('first_name','?')} {p.get('last_name','?')}"
        by_pooler[pooler].append((get_player_bucket(positions.get(pid)), row['player_type'], name, pid))
    for pooler, roster in baseline_roster.items():
        for pid, status in roster.items():
            if (pooler, pid) in sim.rows_by_pair:
                continue
            p = pindex.by_id.get(pid, {})
            name = f"{p.get('first_name','?')} {p.get('last_name','?')}"
            by_pooler[pooler].append((get_player_bucket(positions.get(pid)), status, name, pid))
    return by_pooler


def get_status_before_activation(sim, pooler, player_id, date):
    """Statut qui précédait le passage à 'actif' le plus récent avant `date` — utilisé pour
    savoir à quoi revenir en bannissant un joueur pour une correction de légalité (ex: un
    joueur normalement 'recrue' promu temporairement, comme Fyodor Svechkov/Conor Geekie
    promus le 15 avril, doit redescendre à 'recrue', pas systématiquement 'reserviste')."""
    row = sim.get_open_row(pooler, player_id)
    if row is None:
        return 'reserviste'
    prev_status = None
    for changed_at, old_t, _new_t in sorted(row['transitions'], key=lambda t: t[0]):
        if changed_at <= date:
            prev_status = old_t
    return prev_status or 'reserviste'


def check_legality(sim, baseline_roster, touched_ids, positions, pindex, pooler, date, report):
    roster = get_roster_snapshot(sim, baseline_roster, touched_ids, pooler, date)

    def pname(pid):
        p = pindex.by_id.get(pid, {})
        return f"{p.get('first_name','?')} {p.get('last_name','?')}"

    actifs = [pid for pid, t in roster.items() if t == 'actif']
    reserves = [pid for pid, t in roster.items() if t == 'reserviste']
    by_bucket = {'forward': [], 'defense': [], 'goalie': []}
    for pid in actifs:
        by_bucket[get_player_bucket(positions.get(pid))].append(pid)

    labels = {'forward': "attaquants", 'defense': "défenseurs", 'goalie': "gardiens"}
    for bucket, limit in ACTIVE_LIMITS.items():
        if len(by_bucket[bucket]) > limit:
            names = tuple(sorted(pname(pid) for pid in by_bucket[bucket]))
            msg = f"Trop de {labels[bucket]} actifs ({len(by_bucket[bucket])}/{limit})"
            report.legality_violations.append((date, pooler, bucket, msg, names))
    if len(reserves) < 2:
        names = tuple(sorted(pname(pid) for pid in reserves))
        msg = f"Moins de 2 réservistes ({len(reserves)})"
        report.legality_violations.append((date, pooler, 'reserve', msg, names))


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    apply_mode = '--apply' in sys.argv

    os.makedirs(os.path.join(BASE_DIR, 'logs'), exist_ok=True)
    log_path = os.path.join(BASE_DIR, 'logs', f'import_mouvements_{datetime.now():%Y-%m-%d_%H-%M-%S}.log')
    sys.stdout = Tee(log_path)

    print(f"[INFO] Fichier Excel : {EXCEL_PATH}")
    print(f"[INFO] Mode : {'APPLY' if apply_mode else 'DRY-RUN'}")

    db = connect()
    season = get_active_season(db)
    season_id = season['id']
    season_start_ts = f"{season['saison_start_date']}T12:00:00Z"
    print(f"[INFO] Saison active : {season['season']} (id={season_id}), début={season['saison_start_date']}")

    players = fetch_all(db, 'players', 'id, first_name, last_name, position')
    positions = {p['id']: p['position'] for p in players}
    pindex = PlayerIndex(players)
    print(f"[INFO] {len(players)} joueurs chargés depuis staging.")

    poolers_rows = db.table('poolers').select('id, name').execute().data
    pooler_name_to_id = {p['name']: p['id'] for p in poolers_rows}

    # Baseline : état courant complet, pour les joueurs qui ne seront jamais touchés par
    # le fichier Excel (sert uniquement au calcul de légalité — non modifié par le script).
    current_rosters = fetch_all(db, 'pooler_rosters',
                                 'pooler_id, player_id, player_type, added_at, removed_at',
                                 filters={'pool_season_id': season_id})
    id_to_pooler_name = {p['id']: p['name'] for p in poolers_rows}
    baseline_roster = defaultdict(dict)
    for r in current_rosters:
        if r['removed_at'] is None:
            pname = id_to_pooler_name.get(r['pooler_id'])
            if pname:
                baseline_roster[pname][r['player_id']] = r['player_type']

    excel_rows = load_excel_rows()
    print(f"[INFO] {len(excel_rows)} lignes lues dans la feuille Mouvements.")

    # "Date" est le champ saisi directement par David, toujours renseigné (vérifié : aucune
    # ligne blanche sur les 272). "Date tri" est un champ dérivé (sort_mouvements.py, forward-
    # fill par pooler) calculé une seule fois par le passé — peut désynchroniser de "Date" si
    # "Date" a été corrigée/ajoutée après coup sans relancer ce script (5 lignes concernées,
    # ex: ligne 41 "Date"=2025-12-23 mais "Date tri" encore à 2025-11-08, ce qui faisait
    # traiter la réclamation au ballotage de Morgan Rielly 6 semaines trop tôt et cassait
    # la chronologie de son roster). "Date" prime donc désormais ; "Date tri" ne sert que de
    # repli si "Date" est un jour manquant.
    indexed_rows = list(enumerate(excel_rows, start=2))  # ligne 2 = première ligne de données
    def row_date_key(item):
        _, r = item
        d = to_date_str(r.get('Date')) or to_date_str(r.get('Date tri'))
        return d or '9999-99-99'
    indexed_rows.sort(key=row_date_key)

    sim = Simulation(season_start_ts)
    report = Report()

    # Ensemble complet des joueurs touchés par le fichier, calculé à l'avance (pas construit
    # incrémentalement pendant la boucle) — utilisé pour exclure ces joueurs du baseline dans
    # check_legality. Sans ça, un joueur dont la première mention est tardive (ex: Trevor
    # Zegras, acquis par David seulement le 21 octobre) reste dans le baseline (état DB
    # actuel, potentiellement 'actif') pour toutes les dates AVANT sa première mention — bug
    # trouvé par David en lisant le détail des violations (Zegras apparaissait comme actif
    # chez David dès le 8 octobre, avant même son acquisition réelle).
    all_touched_ids = set()
    for _, r in indexed_rows:
        for cell in (r.get('Joueur acquis/activé'), r.get('Joueur cede/desactive')):
            name = parse_player_name(cell)
            if name:
                pid, _ = pindex.resolve(name)
                if pid is not None:
                    all_touched_ids.add(pid)

    # Roster initial réel (onglet Roster_Initial, ajouté par David le 2026-08-03) — préposé
    # avant de rejouer les mouvements pour remplacer le bootstrap heuristique ('actif' par
    # défaut) par la vraie donnée, pour tout joueur listé.
    roster_initial_rows = load_roster_initial_rows()
    if roster_initial_rows is None:
        print("[INFO] Pas d'onglet 'Roster_Initial' — bootstrap heuristique utilisé pour tous les cas.")
    else:
        seen_players = {}  # player_id -> pooler_name (détection de conflit : 2 poolers pour le même joueur)
        loaded = 0
        for r in roster_initial_rows:
            pooler = map_pooler(r.get('Pooler'))
            if not pooler:
                report.roster_initial_issues.append(f"Pooler non reconnu : {r.get('Pooler')!r} (joueur {r.get('Joueur')!r})")
                continue
            name = parse_player_name(r.get('Joueur'))
            pid, note = pindex.resolve(name) if name else (None, "Nom vide")
            if pid is None:
                report.roster_initial_issues.append(f"{pooler} : {note}")
                continue
            stype = map_status(r.get('Statut'))
            if stype is None or stype == 'BALLOTAGE':
                report.roster_initial_issues.append(
                    f"{pooler} : statut invalide pour {name!r} : {r.get('Statut')!r}")
                continue
            if pid in seen_players:
                report.roster_initial_issues.append(
                    f"{name!r} listé chez {seen_players[pid]} ET {pooler} — deuxième occurrence ignorée")
                continue
            seen_players[pid] = pooler
            sim.preload_initial(pooler, pid, stype)
            loaded += 1
        report.roster_initial_count = loaded
        print(f"[INFO] Roster_Initial : {loaded} joueurs préchargés, {len(report.roster_initial_issues)} problème(s).")

    # Les (date, pooler) à vérifier pour la légalité sont accumulés pendant la boucle
    # principale, mais la vérification elle-même est différée après l'application des
    # corrections manuelles (voir plus bas) — sinon le rapport final ne refléterait pas les
    # bancs déjà corrigés par David. Un seul passage par (date, pooler), pas par ligne
    # individuelle : plusieurs lignes du fichier peuvent partager la même date (plusieurs
    # transactions saisies le même jour) et vérifier après chacune capturerait des états
    # transitoires en plein milieu de journée qui n'ont jamais vraiment existé sur la glace.
    poolers_touched_today = set()
    current_date = None
    date_pooler_pairs = []  # [(date, pooler), ...] dans l'ordre chronologique, dédoublonné

    def flush_legality_checks():
        for p in poolers_touched_today:
            date_pooler_pairs.append((current_date, p))
        poolers_touched_today.clear()

    for idx, r in indexed_rows:
        pooler = map_pooler(r.get('Pooler'))
        if not pooler:
            report.anomalies.append((idx, None, r.get('Pooler'), f"Pooler non reconnu : {r.get('Pooler')!r}"))
            continue
        date = to_date_str(r.get('Date')) or to_date_str(r.get('Date tri'))
        if not date:
            report.anomalies.append((idx, None, pooler, "Aucune date exploitable (Date tri et Date vides)"))
            continue
        if date != current_date:
            if current_date is not None:
                flush_legality_checks()
            current_date = date
        ts = f"{date}T12:00:00Z"
        echange_pooler = map_pooler(r.get('Echange Pooler')) if r.get('Echange Pooler') else None
        report.rows_processed += 1
        report.poolers_touched.add(pooler)
        poolers_touched_today.add(pooler)
        if echange_pooler:
            report.poolers_touched.add(echange_pooler)
            poolers_touched_today.add(echange_pooler)

        # Choix de repêchage — hors scope, listés seulement
        if r.get('Choix acquis') or r.get('Choix cede'):
            report.pick_rows.append((idx, date, pooler,
                f"acquis={r.get('Choix acquis')}/{r.get('Annee choix acquis')} "
                f"cédé={r.get('Choix cede')}/{r.get('Annee choix cede')} "
                f"choix_pooler={r.get('Choix Pooler')}"))

        acquis_name = parse_player_name(r.get('Joueur acquis/activé'))
        if acquis_name:
            pid, note = pindex.resolve(acquis_name)
            if pid is None:
                report.unresolved_names.append((idx, date, pooler, 'acquis', acquis_name, note))
            else:
                report.players_touched.add(pid)
                stype = map_status(r.get('Statut joueur acquis'))
                if stype is None or stype == 'BALLOTAGE':
                    report.anomalies.append((idx, date, pooler,
                        f"Statut acquis invalide/manquant pour {acquis_name} : {r.get('Statut joueur acquis')!r}"))
                else:
                    sim.process_acquis(pooler, pid, ts, stype, echange_pooler, idx, report)

        cede_name = parse_player_name(r.get('Joueur cede/desactive'))
        if cede_name:
            pid, note = pindex.resolve(cede_name)
            if pid is None:
                report.unresolved_names.append((idx, date, pooler, 'cédé', cede_name, note))
            else:
                report.players_touched.add(pid)
                stype = map_status(r.get('Statut joueur cédé'))
                if stype is None:
                    report.anomalies.append((idx, date, pooler,
                        f"Statut cédé invalide/manquant pour {cede_name} : {r.get('Statut joueur cédé')!r}"))
                else:
                    sim.process_cede(pooler, pid, ts, stype, echange_pooler, idx, report)

    flush_legality_checks()  # dernière date du fichier

    # Corrections manuelles de légalité (optionnel, excel/correction_violations_alignements.txt,
    # ajouté par David le 2026-08-04) — appliquées après la simulation Excel complète, avant
    # le rapport final, pour que celui-ci reflète les bancs corrigés plutôt que l'état brut.
    # Pour chaque période (triée chronologiquement par pooler/catégorie), compare la vraie
    # liste d'actifs donnée par David à l'état RÉELLEMENT simulé à cette date précise (pas
    # une supposition basée sur les corrections précédentes) : qui est actif chez nous mais
    # pas chez lui est banni (change_type -> reserviste) ; qui devrait être actif chez lui
    # mais ne l'est pas dans la simulation est réactivé (change_type -> actif). Se base
    # toujours sur un snapshot frais — nécessaire pour bien gérer un joueur réactivé entre-
    # temps par un vrai mouvement du fichier (ex: Ryan Leonard, banni par une correction au
    # 3 novembre, réellement réactivé par le fichier autour du 13, puis re-banni par une
    # autre correction au 17 décembre — un simple suivi "déjà banni par une correction
    # précédente" raterait ce second bannissement puisque la vraie réactivation intermédiaire
    # ne passe pas par le mécanisme de correction). Traité groupe par groupe (indépendants
    # les uns des autres, pas besoin d'un ordre global).
    legality_corrections = load_legality_corrections()
    if legality_corrections:
        for (pooler, bucket), periods in legality_corrections.items():
            for start, corrected_names in periods:
                ts = start  # déjà un timestamp complet ("...T12:00:00Z"), voir load_legality_corrections
                snapshot = get_roster_snapshot(sim, baseline_roster, all_touched_ids, pooler, ts)
                current_active_ids = {pid for pid, t in snapshot.items() if t == 'actif'
                                       and get_player_bucket(positions.get(pid)) == bucket}
                current_names = {}
                for pid in current_active_ids:
                    p = pindex.by_id.get(pid, {})
                    current_names[f"{p.get('first_name','?')} {p.get('last_name','?')}"] = pid

                for name, pid in current_names.items():
                    if name not in corrected_names:
                        bench_type = get_status_before_activation(sim, pooler, pid, ts)
                        sim.change_type(pooler, pid, ts, bench_type, report, f"correction légalité ({bucket})")
                        report.legality_corrections_applied.append((pooler, name, f'banni ({bench_type})', start))

                for name in corrected_names:
                    if name in current_names:
                        continue
                    pid, note = pindex.resolve(name)
                    if pid is None:
                        report.legality_correction_issues.append(
                            f"{pooler} [{start}] ({bucket}) : {note}")
                    elif sim.get_open_row(pooler, pid) is not None:
                        sim.change_type(pooler, pid, ts, 'actif', report, f"réactivation légalité ({bucket})")
                        report.legality_corrections_applied.append((pooler, name, 'réactivé', start))
        print(f"[INFO] Corrections de légalité : {len(legality_corrections)} groupe(s), "
              f"{len(report.legality_corrections_applied)} changement(s) de statut appliqué(s).")
    else:
        print(f"[INFO] Pas de fichier de corrections de légalité ({os.path.basename(CORRECTIONS_PATH)} absent).")

    # Vérification de légalité finale — une fois par (date, pooler), après corrections.
    for date, pooler in date_pooler_pairs:
        check_legality(sim, baseline_roster, all_touched_ids, positions, pindex,
                        pooler, f"{date}T12:00:00Z", report)

    # Corrections manuelles d'alignement final (optionnel, excel/correction_alignements_finaux.txt,
    # ajouté par David le 2026-08-05) — appliquées en dernier, à la date de fin de saison
    # (choix délibéré de garder ça simple plutôt que de demander une date précise par joueur,
    # confirmé avec David). Compare la composition finale voulue à celle réellement simulée ;
    # un même statut dans le même pooler ne fait rien, un statut différent dans le même
    # pooler déclenche un changement de type, un pooler différent déclenche un vrai transfert
    # (fermeture de l'ancienne ligne + ouverture de la nouvelle). Un joueur absent du fichier
    # n'est jamais retiré automatiquement — trop risqué d'inférer une suppression depuis une
    # simple omission ; seulement signalé dans le rapport pour revue manuelle.
    final_alignment_corrections = load_final_alignment_corrections()
    if final_alignment_corrections:
        end_ts = f"{season['saison_end_date']}T12:00:00Z"
        before_rosters = compute_final_rosters(sim, baseline_roster, all_touched_ids, positions, pindex)
        current_by_name = {}
        for pooler, entries in before_rosters.items():
            for _bucket, status, name, pid in entries:
                current_by_name[name] = (pooler, status, pid)

        target_names = {name for d in final_alignment_corrections.values() for name in d}
        for target_pooler, players in final_alignment_corrections.items():
            for name, target_status in players.items():
                pid, note = pindex.resolve(name)
                if pid is None:
                    report.final_alignment_correction_issues.append(f"{target_pooler} : {note}")
                    continue
                cur = current_by_name.get(name)
                cur_pooler, cur_status = (cur[0], cur[1]) if cur else (None, None)
                if cur_pooler == target_pooler and cur_status == target_status:
                    continue
                if cur_pooler == target_pooler:
                    sim.change_type(target_pooler, pid, end_ts, target_status, report, "correction alignement final")
                else:
                    if cur_pooler and sim.get_open_row(cur_pooler, pid) is not None:
                        sim.close_row(cur_pooler, pid, end_ts, report)
                    sim.open_row(target_pooler, pid, end_ts, target_status)
                report.final_alignment_corrections_applied.append(
                    (name, cur_pooler, cur_status, target_pooler, target_status))

        for name, (cur_pooler, cur_status, _pid) in current_by_name.items():
            if name not in target_names:
                report.final_alignment_correction_issues.append(
                    f"{name} présent dans la simulation ({cur_pooler}, {cur_status}) mais absent du "
                    f"fichier de corrections — ignoré, pas retiré automatiquement.")

        print(f"[INFO] Corrections d'alignement final : "
              f"{len(report.final_alignment_corrections_applied)} changement(s) appliqué(s).")
    else:
        print(f"[INFO] Pas de fichier de corrections d'alignement final "
              f"({os.path.basename(FINAL_ALIGNMENT_CORRECTIONS_PATH)} absent).")

    # -----------------------------------------------------------------
    # Rapport
    # -----------------------------------------------------------------
    print("\n" + "=" * 70)
    print("RAPPORT")
    print("=" * 70)
    print(f"Lignes traitées      : {report.rows_processed} / {len(excel_rows)}")
    print(f"Joueurs touchés      : {len(report.players_touched)}")
    print(f"Poolers touchés      : {len(report.poolers_touched)} — {sorted(report.poolers_touched)}")
    print(f"Roster_Initial       : {report.roster_initial_count} joueurs préchargés")

    print(f"\n--- Problèmes Roster_Initial ({len(report.roster_initial_issues)}) ---")
    for msg in report.roster_initial_issues:
        print(f"  {msg}")

    print(f"\n--- Noms non résolus ({len(report.unresolved_names)}) ---")
    for idx, date, pooler, side, name, note in report.unresolved_names:
        print(f"  L{idx} [{date}] {pooler} ({side}) : {note}")

    print(f"\n--- Origines déduites / bootstrap ({len(report.bootstraps)}) ---")
    print("(joueur absent de l'onglet Roster_Initial ET jamais vu avant sa première mention —")
    print(" statut de départ estimé 'actif' par défaut, faute de meilleure info. Si ce nombre")
    print(" est élevé, l'onglet Roster_Initial est peut-être incomplet pour ces joueurs.")
    print(" Triées par écart décroissant depuis le début de saison — plus l'écart est grand,")
    print(" plus le risque de mal attribuer des points sur cette période est élevé. Les")
    print(" premières de la liste méritent une vérification prioritaire ; un écart de 0-2")
    print(" jours a un impact quasi nul peu importe le vrai statut de départ.)")
    season_start_date = datetime.strptime(season['saison_start_date'], '%Y-%m-%d')
    bootstraps_with_gap = []
    for pooler, pid, date, reason in report.bootstraps:
        event_date = datetime.strptime(date[:10], '%Y-%m-%d')
        gap_days = (event_date - season_start_date).days
        bootstraps_with_gap.append((gap_days, pooler, pid, date, reason))
    bootstraps_with_gap.sort(key=lambda t: -t[0])
    for gap_days, pooler, pid, date, reason in bootstraps_with_gap:
        p = pindex.by_id.get(pid, {})
        print(f"  [+{gap_days:>3}j] [{date}] {pooler} <- {p.get('first_name','?')} {p.get('last_name','?')} ({reason})")

    print(f"\n--- Mésaccords 'Echange Pooler' déclaré vs propriétaire simulé ({len(report.echange_mismatches)}) ---")
    for idx, date, pooler, pid, declared, actual in report.echange_mismatches:
        p = pindex.by_id.get(pid, {})
        print(f"  L{idx} [{date}] {pooler} : {p.get('first_name','?')} {p.get('last_name','?')} — Echange Pooler={declared}, propriétaire simulé={actual}")

    print(f"\n--- Anomalies diverses ({len(report.anomalies)}) ---")
    for idx, date, pooler, msg in report.anomalies:
        print(f"  L{idx} [{date}] {pooler} : {msg}")

    print(f"\n--- Lignes avec choix de repêchage, non traitées ({len(report.pick_rows)}) ---")
    for idx, date, pooler, desc in report.pick_rows:
        print(f"  L{idx} [{date}] {pooler} : {desc}")

    print(f"\n--- Corrections de légalité appliquées ({len(report.legality_corrections_applied)}) ---")
    print("(à partir de excel/correction_violations_alignements.txt, si présent)")
    for pooler, name, action, date in sorted(report.legality_corrections_applied, key=lambda t: t[3]):
        print(f"  [{date}] {pooler} : {name} — {action}")

    print(f"\n--- Problèmes dans le fichier de corrections ({len(report.legality_correction_issues)}) ---")
    for msg in report.legality_correction_issues:
        print(f"  {msg}")

    print(f"\n--- Violations de légalité résiduelles, après corrections (non bloquantes, {len(report.legality_violations)}) ---")
    print("(NOTE : le baseline des joueurs jamais touchés par le fichier vient de l'état DB")
    print(" ACTUEL, utilisé comme approximation constante sur toute la saison — un pooler")
    print(" déjà en dépassement aujourd'hui sur ses joueurs non touchés apparaîtra donc en")
    print(" dépassement à chaque ligne le concernant, même si ce n'était pas forcément vrai")
    print(" à l'époque. Signal à interpréter avec prudence, pas une liste de vrais problèmes.)")
    print(" Une ligne par période où la composition fautive reste identique pour un même")
    print(" pooler/catégorie (avec les noms), plutôt qu'une ligne par date traitée — le tout")
    print(" trié par ordre chronologique croissant.")
    by_group = defaultdict(list)  # (pooler, bucket) -> [(date, msg, names), ...]
    for date, pooler, bucket, msg, names in report.legality_violations:
        by_group[(pooler, bucket)].append((date, msg, names))

    runs = []  # (run_start, pooler, label, msg, names)
    for (pooler, bucket), entries in by_group.items():
        entries.sort(key=lambda e: e[0])
        run_start = entries[0][0]
        run_msg, run_names = entries[0][1], entries[0][2]
        prev_date = entries[0][0]
        for date, msg, names in entries[1:] + [(None, None, None)]:
            if names != run_names:
                label = f"du {run_start} au {prev_date}" if run_start != prev_date else run_start
                runs.append((run_start, pooler, label, run_msg, run_names))
                run_start, run_msg, run_names = date, msg, names
            prev_date = date

    runs.sort(key=lambda r: r[0])
    for _, pooler, label, msg, names in runs:
        print(f"  {pooler} [{label}] {msg} : {', '.join(names) if names else '(aucun)'}")

    print(f"\n--- Corrections d'alignement final appliquées ({len(report.final_alignment_corrections_applied)}) ---")
    print("(à partir de excel/correction_alignements_finaux.txt, si présent — appliquées à la date de fin de saison)")
    for name, from_pooler, from_status, to_pooler, to_status in report.final_alignment_corrections_applied:
        print(f"  {name} : {from_pooler}/{from_status} -> {to_pooler}/{to_status}")

    print(f"\n--- Problèmes / signalements alignement final ({len(report.final_alignment_correction_issues)}) ---")
    for msg in report.final_alignment_correction_issues:
        print(f"  {msg}")

    # Diff de sanité : état simulé final vs état actuel réel en base, pour les joueurs touchés
    print(f"\n--- Diff de sanité (simulé final vs DB actuelle, joueurs touchés) ---")
    current_by_player = defaultdict(dict)
    for r in current_rosters:
        if r['removed_at'] is None:
            pname = id_to_pooler_name.get(r['pooler_id'])
            if pname:
                current_by_player[r['player_id']][pname] = r['player_type']

    diffs = 0
    for pid in sorted(report.players_touched):
        sim_final = {}
        for (pooler, ppid), lst in sim.rows_by_pair.items():
            if ppid != pid:
                continue
            row = lst[-1]
            if row['removed_at'] is None:
                sim_final[pooler] = row['player_type']
        db_final = current_by_player.get(pid, {})
        if sim_final != db_final:
            diffs += 1
            p = pindex.by_id.get(pid, {})
            print(f"  {p.get('first_name','?')} {p.get('last_name','?')} — simulé={sim_final} / DB actuelle={db_final}")
    print(f"  ({diffs} joueurs avec un écart sur {len(report.players_touched)} touchés)")

    # Alignements complets : TOUS les joueurs de chaque pooler au terme de la simulation
    # (joueurs touchés par le fichier + joueurs jamais touchés, dont le statut reste celui
    # de la DB actuelle, inchangé) — regroupés par position puis par statut, pour une revue
    # complète pooler par pooler plutôt qu'un simple diff des écarts.
    print(f"\n--- Alignements finaux (tous les joueurs, par position et statut) ---")
    by_pooler_full = compute_final_rosters(sim, baseline_roster, all_touched_ids, positions, pindex)

    bucket_order = {'forward': 0, 'defense': 1, 'goalie': 2}
    bucket_labels = {'forward': 'Attaquants', 'defense': 'Défenseurs', 'goalie': 'Gardiens'}
    status_order = {'actif': 0, 'reserviste': 1, 'ltir': 2, 'recrue': 3}
    for pooler in sorted(by_pooler_full):
        entries = by_pooler_full[pooler]
        print(f"  {pooler} ({len(entries)}) :")
        for bucket in sorted({e[0] for e in entries}, key=lambda b: bucket_order.get(b, 9)):
            bucket_entries = sorted((e for e in entries if e[0] == bucket),
                                     key=lambda e: (status_order.get(e[1], 9), e[2]))
            print(f"    {bucket_labels.get(bucket, bucket)} ({len(bucket_entries)}) :")
            for _, status, name, _pid in bucket_entries:
                print(f"      {status:12s} {name}")

    print("\n" + "=" * 70)
    if not apply_mode:
        print("[DRY-RUN] Aucune écriture effectuée. Relancer avec --apply pour appliquer.")
        return

    # -----------------------------------------------------------------
    # Application réelle
    # -----------------------------------------------------------------
    confirm = input(f"\nAppliquer ces changements à la saison {season['season']} en staging ? (tapez 'oui') ")
    if confirm.strip().lower() != 'oui':
        print("[ANNULÉ] Aucune écriture effectuée.")
        return

    touched_ids = sorted(report.players_touched)
    print(f"[APPLY] Suppression de l'historique existant pour {len(touched_ids)} joueurs...")
    for pid in touched_ids:
        db.table('roster_change_log').delete().eq('pool_season_id', season_id).eq('player_id', pid).execute()
        db.table('pooler_rosters').delete().eq('pool_season_id', season_id).eq('player_id', pid).execute()

    print("[APPLY] Insertion de l'historique simulé...")
    touched_id_set = set(touched_ids)
    inserted_rows, inserted_logs = 0, 0
    for (pooler, pid), lst in sim.rows_by_pair.items():
        if pid not in touched_id_set:
            # Joueur préchargé depuis Roster_Initial mais jamais mentionné dans Mouvements :
            # sa ligne pooler_rosters existante n'a pas été supprimée ci-dessus (scope
            # touched_ids), la réinsérer créerait un doublon (ligne ouverte en double).
            continue
        pooler_id = pooler_name_to_id.get(pooler)
        if not pooler_id:
            print(f"[ERREUR] Pooler inconnu en base : {pooler}")
            continue
        for row in lst:
            db.table('pooler_rosters').insert({
                'pooler_id': pooler_id,
                'player_id': pid,
                'pool_season_id': season_id,
                'player_type': row['player_type'],
                'is_active': row['removed_at'] is None,
                'added_at': row['added_at'],
                'removed_at': row['removed_at'],
            }).execute()
            inserted_rows += 1
            for changed_at, old_type, new_type in row['transitions']:
                db.table('roster_change_log').insert({
                    'player_id': pid, 'pooler_id': pooler_id, 'pool_season_id': season_id,
                    'change_type': 'excel_import', 'old_type': old_type, 'new_type': new_type,
                    'changed_by': None, 'changed_at': changed_at, 'is_admin_override': True,
                }).execute()
                inserted_logs += 1
            if row['removed_at'] is not None:
                db.table('roster_change_log').insert({
                    'player_id': pid, 'pooler_id': pooler_id, 'pool_season_id': season_id,
                    'change_type': 'excel_import', 'old_type': row['player_type'], 'new_type': None,
                    'changed_by': None, 'changed_at': row['removed_at'], 'is_admin_override': True,
                }).execute()
                inserted_logs += 1

    print(f"[APPLY] {inserted_rows} lignes pooler_rosters, {inserted_logs} lignes roster_change_log insérées.")


if __name__ == '__main__':
    main()
